import pandas as pd
import datetime
import time
import pandas_market_calendars as mcal
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from config import etfs_lst, etfs_hedge_date_dict, etfs_calendar_dict, start_date, end_date, excel_path, general_rules_dict

pd.set_option("display.max_rows", 100)
pd.set_option("display.max_columns", 15)
pd.set_option("display.width", 1000)


def get_calendar(mkt_name, start_date, end_date):
    mkt_ex = mcal.get_calendar(mkt_name)
    mkt_ex_schedule = mkt_ex.schedule(start_date=start_date, end_date=end_date)
    mkt_trading_days = mkt_ex_schedule.index.date
    return mkt_trading_days.tolist()


def get_hedge_dict(holding_date, reminder, trading_days):
    t, d = reminder.split('__')
    d0, drift = d.split('+')
    holding_date = datetime.datetime.date(holding_date)
    idx = trading_days.index(holding_date)
    hedge_date = trading_days[idx + int(drift)]
    return hedge_date, t


def match_trading_days_from_reminders(holding_date, Fixing_reminder, FX_reminder, trading_days, etf_hedge_num_ser, general_rules_dataframe):
    holding_date_str = str(holding_date.year) + '-' + str(holding_date.month) + '-' + str(holding_date.day)
    hedge_dict = dict()
    etf_hedge_num_df_temp = etf_hedge_num_ser
    source = 'source:' + holding_date_str + '_' + str(etf_hedge_num_df_temp.loc['etf_code']) + '_' + \
             str(etf_hedge_num_df_temp.loc['C/R']) + '_' + str(etf_hedge_num_df_temp.loc['UNDERLYING']) + '_' + \
             str(etf_hedge_num_df_temp.loc['SETTEL CURRENCY']) + '_' + str(etf_hedge_num_df_temp.loc['NUM'])
    # hedge的资产和量
    underlying = etf_hedge_num_df_temp.loc['UNDERLYING']
    num = etf_hedge_num_df_temp.loc['NUM']
    market = etf_hedge_num_df_temp.loc['MARKET']

    i = 1
    if market == 'INDEX':
        if ',' in Fixing_reminder:
            pass
        else:
            k1, v1 = get_hedge_dict(holding_date, Fixing_reminder, trading_days)
            general_rule = general_rules_dataframe.loc[str(etf_hedge_num_df_temp.loc['etf_code']), 'Fixing']
            specific_marker = '' if Fixing_reminder == general_rule else '_specific'
            hedge_dict[i] = 'handle_date:' + str(k1)
            hedge_dict[i] += ';' + 'handle_time:' + v1.split('_')[-1]
            hedge_dict[i] += ';' + source + '_' + Fixing_reminder.replace("_", "") + specific_marker
            hedge_dict[i] += ';' + 'num:' + str(num)
            hedge_dict[i] += ';' + 'underlying:' + underlying

    elif market == 'FX':
        if ',' in FX_reminder:
            reminder_lst = FX_reminder.split(',')
            general_rule = general_rules_dataframe.loc[str(etf_hedge_num_df_temp.loc['etf_code']), 'FX']
            general_rule_lst = general_rule.split(',')
            for reminder, rule in zip(reminder_lst, general_rule_lst):
                path = etf_hedge_num_df_temp.loc['PATH']
                if path in reminder:
                    ki, vi = get_hedge_dict(holding_date, reminder, trading_days)
                    specific_marker = '' if reminder == rule else '_specific'
                    hedge_dict[i] = 'handle_date:' + str(ki)
                    hedge_dict[i] += ';' + 'handle_time:' + vi.split('_')[-1]
                    hedge_dict[i] += ';' + source + '_' + reminder.replace("_", "") + specific_marker
                    hedge_dict[i] += ';' + 'num:' + str(num)
                    hedge_dict[i] += ';' + 'underlying:' + underlying
        else:
            k2, v2 = get_hedge_dict(holding_date, FX_reminder, trading_days)
            general_rule = general_rules_dataframe.loc[str(etf_hedge_num_df_temp.loc['etf_code']), 'Fixing']
            specific_marker = '' if Fixing_reminder == general_rule else '_specific'
            hedge_dict[i] = 'handle_date:' + str(k2)
            hedge_dict[i] += ';' + 'handle_time:' + v2.split('_')[-1]
            hedge_dict[i] += ';' + source + '_' + FX_reminder.replace("_", "") + specific_marker
            hedge_dict[i] += ';' + 'num:' + str(num)
            hedge_dict[i] += ';' + 'underlying:' + underlying

    return hedge_dict


def get_etf_trading_days(etf_code, start_date, end_date, etfs_calendar_dict):
    opendate_mkt_lst = etfs_calendar_dict[str(etf_code)]
    if len(opendate_mkt_lst) == 1:
        trading_days = get_calendar(opendate_mkt_lst[0], start_date, end_date)
    if len(opendate_mkt_lst) > 1:
        trading_days = set()
        for i in opendate_mkt_lst:
            mkt_i_trading_days = get_calendar(i, start_date, end_date)
            trading_days = trading_days & set(mkt_i_trading_days) if len(trading_days) != 0 else set(mkt_i_trading_days)
        trading_days = sorted(list(trading_days))
    return trading_days


def match_Fixing_FX(etfs_lst, etfs_holding_dataframe, etfs_hedge_date_dataframe, general_rules_dataframe, start_date, end_date, etfs_calendar_dict):

    hedge_df = pd.DataFrame()
    for iidx in range(etfs_holding_dataframe.shape[0]):
        etf_hedge_num_ser = etfs_holding_dataframe.iloc[iidx, :]
        etf_code = etf_hedge_num_ser.loc['etf_code']
        date = etf_hedge_num_ser.loc['Date']
        Fixing_reminder = etfs_hedge_date_dataframe.loc[str(etf_code), 'Fixing']
        FX_reminder = etfs_hedge_date_dataframe.loc[str(etf_code), 'FX']
        trading_days = get_etf_trading_days(etf_code, start_date, end_date, etfs_calendar_dict)
        hedge_dict = match_trading_days_from_reminders(date, Fixing_reminder, FX_reminder, trading_days, etf_hedge_num_ser, general_rules_dataframe)
        hedge_df.loc[iidx, 'info'] = hedge_dict[1]

    return hedge_df


def cal_repetitive_info(df):
    df_repetitive = df[[',' in x for x in df['info']]]
    df_no_repetitive = df[[',' not in x for x in df['info']]]
    info_repetitive_df = pd.DataFrame()
    if len(df_repetitive):
        for idx in df_repetitive.index:
            info_repetitive = df_repetitive.loc[idx, 'info']
            info_repetitive_lst = info_repetitive.split(',')
            for i, info in enumerate(info_repetitive_lst):
                info_repetitive_df.loc[i, 'info'] = info
    df_trans = pd.concat([info_repetitive_df, df_no_repetitive], ignore_index=True)

    df_trans['handle_date'] = df_trans['info'].map(lambda x: x.split(';')[0].split(':')[1])
    df_trans['handle_time'] = df_trans['info'].map(lambda x: ':'.join(x.split(';')[1].split(':')[1:]).split('_')[-1])
    df_trans['source'] = df_trans['info'].map(lambda x: ':'.join(x.split(';')[2].split(':')[1:]))
    df_trans['num'] = df_trans['info'].map(lambda x: float(x.split(';')[3].split(':')[1]))
    df_trans['underlying'] = df_trans['info'].map(lambda x: x.split(';')[4].split(':')[1])
    df_trans.drop(['info'], axis=1, inplace=True)
    df_trans = df_trans[pd.to_datetime(df_trans['handle_date']).dt.date >= datetime.date.today()]  # 实际上只输出当天，及之后的信息
    df_trans = df_trans.sort_values(by=['handle_date', 'handle_time']).reset_index(drop=True)
    df_trans = df_trans.reindex(columns=['handle_date', 'handle_time', 'num', 'underlying', 'source'])

    return df_trans


def automation(etfs_lst, etfs_hedge_date_dict, general_rules_dict, excel_path, start_date, end_date, etfs_calendar_dict):
    etfs_holding_dataframe = pd.read_excel(excel_path, sheet_name=0)
    etfs_hedge_date_dataframe = pd.DataFrame(etfs_hedge_date_dict).T
    general_rules_dataframe = pd.DataFrame(general_rules_dict).T
    hedge_df = match_Fixing_FX(etfs_lst, etfs_holding_dataframe, etfs_hedge_date_dataframe, general_rules_dataframe, start_date, end_date, etfs_calendar_dict)
    hedge_df_trans = cal_repetitive_info(hedge_df)

    wb = load_workbook(filename=excel_path)
    if 'hedge_info' in wb.sheetnames:
        ws1 = wb['hedge_info']
        wb.remove(ws1)
    wb.save(excel_path)
    wb.close()

    with pd.ExcelWriter(excel_path, mode='a') as writer:
        hedge_df_trans.to_excel(writer, sheet_name='hedge_info')

    # 对specific_rule填充颜色
    workbook = load_workbook(filename=excel_path)
    sheet = workbook['hedge_info']
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for cell in sheet['F']:
        if 'specific' in str(cell.value):
            cell.fill = fill

    workbook.save(excel_path)
    workbook.close()


if __name__ == '__main__':
    start_time = time.time()
    automation(etfs_lst, etfs_hedge_date_dict, general_rules_dict, excel_path, start_date, end_date, etfs_calendar_dict)
    print("The calculation is complete: {} s".format(round(time.time() - start_time, 2)))
